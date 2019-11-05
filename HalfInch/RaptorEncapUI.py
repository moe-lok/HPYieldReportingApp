import tkinter as tk
from tkinter import StringVar, messagebox, BOTH, LEFT, BOTTOM, RIGHT, TOP, X, END
from tkinter.ttk import Frame, LabelFrame, Label, Entry, Button
import datetime
from copy import copy

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from tkcalendar import DateEntry


class RaptorEncap(tk.Frame):
    # Column name of table
    fieldsLeft = ('Operator', 'Reel ID', 'Qty In', 'Qty Out')
    fieldsRight = ('Bead Height', 'Total Bead Void', 'Big Head', 'Large Bead', 'Small Bead', 'Insufficient Encap',
                   'Prodn Scrap', 'Others', 'Remarks')

    @staticmethod
    def resetForm(entries):
        for k, entry in entries.items():
            print(k + " " + entry.get())
            entry.delete(0, END)

    @staticmethod
    def resetValid(lbl):
        print("reset valid ######")
        for v in lbl:
            print("before reset " + str(lbl[v].grid_info()))
            lbl[v].grid_remove()
            print("after reset " + str(lbl[v].grid_info()))

    @staticmethod
    def makeForm(parent, fields):
        entries = {}
        for field in fields:
            row = Frame(parent)
            lab = Label(row, text=field + ": ", anchor='w')
            ent = Entry(row)
            row.pack(side=TOP, fill=X, padx=5, pady=5)
            lab.pack(side=LEFT)
            ent.pack(side=RIGHT)
            entries[field] = ent
        return entries

    @staticmethod
    def makeValidateTxt(parent, fields):
        labels = {}
        print("validate txt ########")
        for idx, val in enumerate(fields):
            row = Frame(parent)
            row.pack()
            lab = Label(row, text=val + " is empty ", foreground="red")
            lab.grid(row=idx, column=0, padx=5, pady=6)
            print("before remove " + str(lab.grid_info()['row']))
            lab.grid_remove()
            print("after remove " + str(lab.grid_info()))
            labels[val] = lab
        return labels

        # validation callback function

    @staticmethod
    def onValidate(inputStr):
        if inputStr.isdigit():
            print("isidigit")
            return True
        elif inputStr.strip() == "":
            return False
        else:
            return False

    def __init__(self, root, strFileDir):
        tk.Frame.__init__(self, root)

        SHEET_NAME = "RT Encap"

        TarCoverlayerYieldTarget = 99.4

        colDict = {
            "colOperator": "D",
            "colReelID": "E",
            "colQtyIn": "F",
            "colQtyOut": "G",

            # Reject Code
            "colBeadHeight": "I",
            "colTotalBeadVoid": "J",
            "colBigHead": "K",
            "colLargeBead": "L",
            "colSmallBead": "M",
            "colInsufficientEncap": "N",
            "colProdnScrap": "O",
            "colOthers": "P",
            "colRemarks": "Q"
        }

        colDict2 = {
            "colDate": "B",
            "colDay": "C",
            "colYieldTarget": "H",
        }

        # String var for realtime trigger
        strQtyIn = StringVar()
        strQtyOut = StringVar()
        strQtyIn.trace("w", lambda name, index, mode, sv=strQtyIn: calculateYield())
        strQtyOut.trace("w", lambda name, index, mode, sv=strQtyOut: calculateYield())

        def calculateYield():
            val1 = strQtyIn.get()
            val2 = strQtyOut.get()
            try:
                result = int(val2 if val2 else 0) / int(val1 if val1 else 0) * 100
                lblYieldTarget['text'] = round(result, 2)
                if round(result, 2) < TarCoverlayerYieldTarget:
                    lblYieldTarget.config(background="red")
                else:
                    lblYieldTarget.config(background="green")
            except ZeroDivisionError:
                lblYieldTarget['text'] = "fill up next value"

        def resetForms():
            self.resetForm(ents)
            self.resetValid(valid)
            validMainFrame.pack_forget()
            validRejFrame.pack_forget()

        def checkForEmpty(entries, valids):
            print("Check for empty entries****")
            isEmpty = False
            self.resetValid(valids)
            validMainFrame.pack_forget()
            validRejFrame.pack_forget()

            for k, entry in entries.items():
                if entry.get().strip() == "":
                    validMainFrame.pack()
                    validRejFrame.pack()
                    print("empty " + k)
                    isEmpty = True
                    print("before empty " + str(valids[k].grid_info()))
                    valids[k].grid()
                    print("after empty " + str(valids[k].grid_info()['row']))

            if not isEmpty:
                print("not empty")

            return isEmpty

        def copyStylePrevRow(ws, cell, nextRow):
            print("copy style from previous row****")
            # copy style from previous row
            new_cell = ws[cell.column_letter + nextRow]
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

        def copyFormulaPrevRow(ws, lastItem):
            print("copy formula from previous row****")
            # H
            # iterate previous row
            for cell in ws["{}{}:{}{}".format(colDict2["colDate"], lastItem[0].row,
                                              colDict["colRemarks"], lastItem[0].row)][0]:
                print(cell.value)
                prevRow = str(cell.row)
                nextRow = str(cell.row + 1)
                # copy formula from previous row
                ws[colDict2["colDate"] + nextRow] = ws[colDict2["colDate"] + prevRow].value
                ws[colDict2["colDay"] + nextRow] = ws[colDict2["colDay"] + prevRow].value

                ws[colDict2["colYieldTarget"] + nextRow] = Translator(ws[colDict2["colYieldTarget"] + prevRow].value,
                                                                      origin=colDict2["colYieldTarget"] + prevRow) \
                    .translate_formula(colDict2["colYieldTarget"] + nextRow)

                # copy style from previous row
                copyStylePrevRow(ws, cell, nextRow)

        def searchRowDate(ws, d):
            print("Search for ros=w of matching date****")
            # get length of row
            row_count = len(ws[colDict2["colDate"]])

            # row_count = ws1.max_row
            print("row count is: " + str(row_count))

            # TODO: adjust minimum
            # set the minimum and maximum
            minRow = 7
            maxRow = row_count - 1
            dateRange = ws['{}{}:{}{}'.format(colDict2["colDate"], minRow, colDict2["colDate"], maxRow)]
            lastItem = any
            b1 = datetime.date(d.year, d.month, d.day)

            for cell in dateRange:
                print(cell[0].row, cell[0].value)
                c = cell[0].value
                try:

                    b2 = datetime.date(c.year, c.month, c.day)
                    if b1 == b2:
                        print("same date")
                        lastItem = cell
                except AttributeError as e:
                    print(e)

            return lastItem

        def fillCell(ws, lastItem, offset):
            print("Fill up cell****")
            # fill up appropriate cell
            rowOffset = str(lastItem[0].row + offset)  # get row plus offset
            #
            for idx, val in enumerate(ents):
                if val in ents1:
                    print(val)
                    ws[colDict[list(colDict)[idx]] + rowOffset] = ents[val].get()

                if val in ents2:
                    if val == 'Remarks':
                        print(val)
                        ws[colDict[list(colDict)[idx]] + rowOffset] = ents[val].get() if ents[val].get() else None
                    else:
                        print(val)
                        ws[colDict[list(colDict)[idx]] + rowOffset] = int(ents[val].get()) if ents[
                            val].get() else None

        def modifyRow(ws, lastItem):
            print("Modify Row****")
            # check whether cell to fill in are empty
            if ws['{}{}'.format(colDict["colReelID"], str(lastItem[0].row))].value is None:
                # if empty, fill up appropriate cell
                fillCell(ws, lastItem, 0)
            else:
                # get length of row
                row_count = len(ws[colDict2["colDate"]])
                # if not, insert new row below by moving the cells one row down
                ws.move_range("{}{}:{}{}".format(colDict2["colDate"], lastItem[0].row + 1, colDict["colRemarks"],
                                                 row_count), rows=1, translate=True)
                copyFormulaPrevRow(ws, lastItem)
                # then, fill up appropriate cell
                fillCell(ws, lastItem, 1)

        def handle_submit():
            # submit only when required fields are filled
            if not checkForEmpty(entries2, valid):
                # double confirm submit
                if messagebox.askokcancel('Submit', 'Are you sure you want to submit?'):
                    print("Ready to submit****")
                    file_location = strFileDir.get()

                    try:  # load workbook
                        wb = load_workbook(filename=file_location)
                    except PermissionError:
                        messagebox.showerror("Fail to load", "Permission Error:\n"
                                                             "User does not have permission to access or\n"
                                                             "Workbook is opened elsewhere")
                    else:
                        # get appropriate worksheet
                        ws = wb[SHEET_NAME]
                        # get date
                        date = entryCal.get_date()
                        # get row based on date
                        lastItem = searchRowDate(ws, date)
                        # check if empty and modify accordingly
                        modifyRow(ws, lastItem)
                        try:
                            wb.save(filename=file_location)
                        except PermissionError:
                            messagebox.showerror("Fail to save", "Permission Error:\n"
                                                                 "User does not have permission to access or\n"
                                                                 "Workbook is opened elsewhere")
                        except Exception as e:
                            print(e)
                            messagebox.showerror("Fail to save", "something went wrong when saving the changes\n"
                                                                 "your changes has not been submitted\n" + str(e))
                        else:
                            self.resetForm(ents)
                            messagebox.showinfo("Submitted", "Entry successfully submitted")

        # Frames

        # Top Frame
        topFrame = tk.Frame(root)
        topFrame.pack(fill=BOTH, expand=True)

        # Top Left Frame
        leftFrame = tk.Frame(topFrame)
        leftFrame.pack(side=LEFT, fill=BOTH, padx=10, pady=10)

        mainFrame = tk.Frame(leftFrame)
        mainFrame.pack(side=LEFT, fill=BOTH)

        validMainFrame = tk.Frame(leftFrame, pady=40)
        validMainFrame.pack(side=LEFT, fill=BOTH, expand=True)
        validMainFrame.pack_forget()

        # Top Right Frame
        rightFrame = LabelFrame(topFrame, text='Reject Code')
        rightFrame.pack(padx=10, pady=10)

        rejFrame = Frame(rightFrame)
        rejFrame.pack(side=LEFT, fill=BOTH, expand=True)

        validRejFrame = Frame(rightFrame)
        validRejFrame.pack(side=LEFT, fill=BOTH, expand=True)
        validRejFrame.pack_forget()

        # bottom frame
        bottomFrame = tk.Frame(root)
        bottomFrame.pack(side=BOTTOM, fill=BOTH, expand=True, padx=10, pady=10)

        # date input
        dateRow = Frame(mainFrame)
        dateRow.pack(side=TOP, fill=X, padx=5, pady=5)
        lblDate = Label(dateRow, text="Date:", anchor='w')
        lblDate.pack(side=LEFT)
        entryCal = DateEntry(dateRow, width=12, background='blue',
                             foreground='white', borderwidth=2)
        entryCal.pack(side=RIGHT, padx=10)

        # making the rest of the form
        ents = {}
        ents1 = self.makeForm(mainFrame, self.fieldsLeft)
        entries2 = ents1

        ents2=self.makeForm(rejFrame, self.fieldsRight)

        ents.update(ents1)
        ents.update(ents2)

        valid = self.makeValidateTxt(validMainFrame, self.fieldsLeft)

        lblYieldTarget = Label(mainFrame, text="yield")
        lblYieldTarget.pack(padx=5, pady=5)

        reg = root.register(self.onValidate)

        # validation to only allow numbers
        ents['Qty In'].config(textvariable=strQtyIn, validate="key", validatecommand=(reg, '%S'))
        ents['Qty Out'].config(textvariable=strQtyOut, validate="key", validatecommand=(reg, '%S'))

        # buttons
        btnReset = Button(bottomFrame, text="reset", width=12, command=resetForms).pack(side=LEFT, pady=5, padx=5, )
        btnSubmit = Button(bottomFrame, text="submit", width=12, command=handle_submit).pack(side=LEFT, pady=5,
                                                                                             padx=5, )

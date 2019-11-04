import tkinter as tk
from tkinter import StringVar, messagebox, BOTH, LEFT, BOTTOM, RIGHT, TOP, X, END, RAISED
from tkinter.ttk import Frame, LabelFrame, Label, Entry, Button
import datetime
from copy import copy

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from tkcalendar import DateEntry


class RaptorBonder(tk.Frame):
    # Column name of table
    fieldsLeft = ('Operator', 'THA Reel ID')
    fieldsLeft1 = ('Wafer Lot #', 'Wafer ID #')
    fieldsLeft2 = ('Oriflex Lot #', 'Die Bond', 'THA Actual Out')
    fieldsRight1 = ('Die In', 'M/C Reject')
    fieldsRight2 = ('Unreadable FIDs', 'Flex For Tape Feed', 'Flex For Audit Test', 'High Alignment', 'Other')
    fieldsRight3 = ('THA Flush Test', 'Shear Test', 'E-test', 'Others', 'Remarks')
    rows = 1
    multiWidget = []

    def resetForm(self, entries):
        for k, entry in entries.items():
            if entry.winfo_class() != "TFrame":
                print(k + " " + entry.get())
                entry.delete(0, END)

        length = len(self.multiWidget)
        for i in range(length, 1, -1):
            self.delField(self.multiWidget)

    @staticmethod
    def resetValid(lbl):
        print("reset valid ######")
        for v in lbl:
            lbl[v].grid_remove()
            print("after reset ")

    @staticmethod
    def makeForm(parent, fields):
        entries = {}
        for field in fields:
            row = Frame(parent)
            lab = Label(row, text=field + ": ", anchor='w')
            ent = Entry(row)
            row.pack(side=TOP, fill=X, padx=5, pady=5)
            lab.pack(side=LEFT)
            ent.pack(side=RIGHT)
            entries[field] = ent
        return entries

    @staticmethod
    def makeValidateTxt(parent, fields):
        labels = {}
        print("validate txt ########")
        for idx, val in enumerate(fields):
            row = Frame(parent)
            row.pack()
            lab = Label(row, text=val + " is empty ", foreground="red")
            lab.grid(row=idx, column=0, padx=5, pady=6)
            print("before remove " + str(lab.grid_info()['row']))
            lab.grid_remove()
            print("after remove " + str(lab.grid_info()))
            labels[val] = lab
        return labels

        # validation callback function

    @staticmethod
    def onValidate(inputStr):
        if inputStr.isdigit():
            print("isidigit")
            return True
        elif inputStr.strip() == "":
            return False
        else:
            return False

    @staticmethod
    def addField(parentFrame, multiWidget):
        multi = Frame(parentFrame, relief=RAISED)
        multi.pack(side=TOP, fill=X, padx=5, pady=5)
        row = Frame(multi)
        row2 = Frame(multi)
        lab = Label(row, text='Wafer Lot #' + ": ", anchor='w')
        ent = Entry(row)
        lab1 = Label(row2, text='Wafer ID #' + ": ", anchor='w')
        ent1 = Entry(row2)
        row.pack(side=TOP, fill=X, padx=5, pady=5)
        row2.pack(side=TOP, fill=X, padx=5, pady=5)
        lab.pack(side=LEFT)
        ent.pack(side=RIGHT)
        lab1.pack(side=LEFT)
        ent1.pack(side=RIGHT)

        multiWidget.append({"Wafer Lot #": ent, "Wafer ID #": ent1, "multi": multi})

    @staticmethod
    def delField(multiWidget):
        multiWidget[-1]["Wafer Lot #"].destroy()
        multiWidget[-1]["Wafer ID #"].destroy()
        multiWidget[-1]["multi"].destroy()
        multiWidget.pop()

    def __init__(self, root, strFileDir):
        tk.Frame.__init__(self, root)

        SHEET_NAME = "RT Bonder"

        RapBonderYieldTarget = 98.75

        colDict = {
            "colOperator": "D",
            "colTHAReelID": "E",
            "colWaferLot": "F",
            "colWaferID": "G",
            "colOriflexLot": "H",
            "colDieBond": "I",
            "colTHAActualOut": "J",

            # Dies
            "colDieIn": "L",
            "colMCReject": "M",

            # Flex
            "colUnreadableFIDs": "O",
            "colFlexForTapeFeed": "P",
            "colFlexForAuditTest": "Q",
            "colHighAlignment": "R",
            "colOther": "S",

            # THA
            "colTHAFlushTest": "U",
            "colShearTest": "V",
            "colEtest": "W",
            "colOthers": "X",
            "colRemarks": "Z"
        }

        colDict2 = {
            "colDate": "B",
            "colDay": "C"
        }

        colDict3 = {
            "colYieldTarget": "K",
            "colRejDies": "N",
            "colSetupAuditFlex": "T",
            "colSetupAuditTHA": "Y"
        }

        # String var for realtime trigger
        strQtyIn = StringVar()
        strQtyOut = StringVar()
        strQtyIn.trace("w", lambda name, index, mode, sv=strQtyIn: calculateYield())
        strQtyOut.trace("w", lambda name, index, mode, sv=strQtyOut: calculateYield())

        def calculateYield():
            val1 = strQtyIn.get()
            val2 = strQtyOut.get()
            try:
                result = int(val2 if val2 else 0) / int(val1 if val1 else 0) * 100
                lblYieldTarget['text'] = round(result, 2)
                if round(result, 2) < RapBonderYieldTarget:
                    lblYieldTarget.config(background="red")
                else:
                    lblYieldTarget.config(background="green")
            except ZeroDivisionError:
                lblYieldTarget['text'] = "fill up next value"

        def resetForms():
            self.resetForm(ents)
            self.resetValid(valid)
            validMainFrame.pack_forget()
            validDiesMainFrame.pack_forget()
            validFlexMainFrame.pack_forget()
            validThaMainFrame.pack_forget()

        def checkForEmpty(entries, valids):
            print("Check for empty entries****")
            isEmpty = False
            self.resetValid(valids)
            validMainFrame.pack_forget()
            validDiesMainFrame.pack_forget()
            validFlexMainFrame.pack_forget()
            validThaMainFrame.pack_forget()

            for k, entry in entries.items():
                if entry.get().strip() == "":
                    validMainFrame.pack()
                    validDiesMainFrame.pack()
                    validFlexMainFrame.pack()
                    validThaMainFrame.pack()
                    print("empty " + k)
                    isEmpty = True
                    print("before empty " + str(valids[k].grid_info()))
                    valids[k].grid()
                    print("after empty " + str(valids[k].grid_info()['row']))

            if not isEmpty:
                print("not empty")

            return isEmpty

        def copyStylePrevRow(ws, cell, nextRow):
            print("copy style from previous row****")
            # copy style from previous row
            new_cell = ws[coordinate_from_string(cell.coordinate)[0] + nextRow]
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

        def copyFormulaPrevRow(ws, lastItem):
            print("copy formula from previous row****")
            # H
            # iterate previous row

            for cell in ws["{}{}:{}{}".format(colDict2["colDate"], lastItem.row,
                                              colDict["colRemarks"], lastItem.row)][0]:
                print(cell.value)
                prevRow = cell.row
                currRow = str(cell.row + 1)
                # copy formula from previous row
                ws[colDict2["colDate"] + currRow] = ws[nextIfNone(ws, colDict2["colDate"], prevRow)].value.date()
                ws[colDict2["colDate"] + currRow].number_format = 'D-MMM-YY'
                ws[colDict2["colDay"] + currRow] = ws[nextIfNone(ws, colDict2["colDay"], prevRow)].value

                for column in colDict3.values():
                    ws[column + currRow] = Translator(ws[nextIfNone(ws, column, prevRow)].value, origin=nextIfNone(
                        ws, column, prevRow)).translate_formula(column + currRow)

                # copy style from previous row
                copyStylePrevRow(ws, cell, currRow)

        def nextIfNone(ws, column, row):
            if ws[column + str(row)].value is None:
                return nextIfNone(ws, column, row - 1)
            else:
                return column + str(row)

        def nextIfMerged(ws, cell, row):
            print("inside next is merged ####")
            if type(ws[cell[0].column_letter + str(row + 1)]).__name__ == 'MergedCell':
                print(ws[cell[0].column_letter + str(row + 1)])
                return nextIfMerged(ws, cell, row + 1)
            else:
                print("next row is not none")
                print(ws[cell[0].column_letter + str(row)])
                return ws[cell[0].column_letter + str(row)]

        def searchRowDate(ws, d):
            print("Search for ros=w of matching date****")
            # get length of row
            row_count = len(ws[colDict2["colDate"]])

            # row_count = ws1.max_row
            print("row count is: " + str(row_count))

            # TODO: adjust minimum
            # set the minimum and maximum
            minRow = 7
            maxRow = row_count - 1
            dateRange = ws['{}{}:{}{}'.format(colDict2["colDate"], minRow, colDict2["colDate"], maxRow)]
            lastItem = any
            b1 = datetime.date(d.year, d.month, d.day)

            for row, cell in enumerate(dateRange, minRow):
                print(cell[0].row, cell[0].value)
                c = cell[0].value
                try:

                    b2 = datetime.date(c.year, c.month, c.day)
                    if b1 == b2:
                        print("same date")
                        print(cell)
                        print(cell[0].column_letter)
                        print(str(row))
                        print(ws[cell[0].column_letter + str(row)])
                        # If next row of date is None means its a merged cell
                        lastItem = nextIfMerged(ws, cell, row)
                        print(lastItem)
                except AttributeError as e:
                    print(e)

            return lastItem

        def fillCell(ws, lastItem, offset, keys):
            print("Fill up cell****")
            # fill up appropriate cell
            rowOffset = str(lastItem.row + offset)  # get row plus offset
            #
            print(keys)
            for idx, val in enumerate(keys):

                if val in {**ents3, **ents4, **ents5}:
                    if val == 'Remarks':
                        print(val)
                        ws[colDict[list(colDict)[idx]] + rowOffset] = ents[val].get() if ents[val].get() else None
                    else:
                        print(val)
                        ws[colDict[list(colDict)[idx]] + rowOffset] = int(ents[val].get()) if ents[
                            val].get() else None
                else:
                    print(val)
                    ws[colDict[list(colDict)[idx]] + rowOffset] = ents[val].get()

        def modifyRow(ws, lastItem):
            print("Modify Row****")

            thaCell = ws['{}{}'.format(colDict["colTHAReelID"], str(lastItem.row))]

            # check whether cell to fill in are empty
            if type(thaCell).__name__ == 'MergedCell' or not (thaCell.value is None):

                if len(self.multiWidget) <= 1:
                    print("multiwidget only 1 #######")

                    # get length of row
                    row_count = len(ws[colDict2["colDate"]])
                    # if not, insert new row below by moving the cells one row down
                    ws.move_range("{}{}:{}{}".format(colDict2["colDate"], lastItem.row + 1, colDict["colRemarks"],
                                                     row_count), rows=1, translate=True)
                    copyFormulaPrevRow(ws, lastItem)
                    # then, fill up appropriate cell
                    fillCell(ws, lastItem, 1, keys)

                else:
                    print("multiwidget more than 1 #######")
                    moveRows = len(self.multiWidget)
                    # get length of row
                    row_count = len(ws[colDict2["colDate"]])
                    # if not, insert new row below by moving the cells one row down
                    ws.move_range("{}{}:{}{}".format(colDict2["colDate"], lastItem.row + 1, colDict["colRemarks"],
                                                     row_count), rows=moveRows, translate=True)
                    copyFormulaPrevRow(ws, lastItem)
                    fillMulti(ws, lastItem, self.multiWidget, keys, 1)
                    fillCell(ws, lastItem, 1, keys)
                    mergeCell(ws, lastItem, 1)

                    # TODO: fill wafer lot # and wafer id #
                    # TODO: fill the rest of column
                    # TODO: merge rows of certain column

            else:
                # if empty, fill up appropriate cell
                if len(self.multiWidget) <= 1:

                    # then, fill up appropriate cell
                    fillCell(ws, lastItem, 0, keys)
                else:
                    print("multiwidget more than 1 #######")
                    moveRows = len(self.multiWidget)
                    # get length of row
                    row_count = len(ws[colDict2["colDate"]])
                    # if not, insert new row below by moving the cells one row down
                    ws.move_range("{}{}:{}{}".format(colDict2["colDate"], lastItem.row + 1, colDict["colRemarks"],
                                                     row_count), rows=moveRows - 1, translate=True)
                    copyFormulaPrevRow(ws, lastItem)
                    fillMulti(ws, lastItem, self.multiWidget, keys, 0)
                    fillCell(ws, lastItem, 0, keys)
                    mergeCell(ws, lastItem, 0)

        def mergeCell(ws, lastItem, offset):
            rowOffset = lastItem.row + offset  # get row plus offset
            print("inside mergeCell #######")
            print(keys)
            for idx, val in enumerate(keys):

                if val in {**ents1, **ents2, **ents3, **ents4, **ents5}:
                    print(val)
                    mergeCol = colDict[list(colDict)[idx]]
                    print("{}{}:{}{}".format(mergeCol, rowOffset, mergeCol, rowOffset + len(self.multiWidget) - 1))
                    ws.merge_cells(
                        "{}{}:{}{}".format(mergeCol, rowOffset, mergeCol, rowOffset + len(self.multiWidget) - 1))

            for val in {**colDict2, **colDict3}.values():
                print(val)
                print("{}{}:{}{}".format(val, rowOffset, val, rowOffset + len(self.multiWidget) - 1))
                ws.merge_cells("{}{}:{}{}".format(val, rowOffset, val, rowOffset + len(self.multiWidget) - 1))

        def fillMulti(ws, lastItem, multi, keys, offset):
            print("inside fillMulti #######")
            rowOffset = lastItem.row + offset  # get row plus offset

            for idx, w in enumerate(multi):
                print(idx)
                print(w["Wafer Lot #"].get())
                print(w["Wafer ID #"].get())

                ws[colDict[list(colDict)[keys.index("Wafer Lot #")]] + str(rowOffset + idx)] = w["Wafer Lot #"].get()
                ws[colDict[list(colDict)[keys.index("Wafer ID #")]] + str(rowOffset + idx)] = w["Wafer ID #"].get()

        def handle_submit():
            # submit only when required fields are filled
            if not checkForEmpty(entries2, valid):

                # double confirm submit
                if messagebox.askokcancel('Submit', 'Are you sure you want to submit?'):
                    print("Ready to submit****")
                    file_location = strFileDir.get()

                    try:  # load workbook
                        wb = load_workbook(filename=file_location)
                    except PermissionError:
                        messagebox.showerror("Fail to load", "Permission Error:\n"
                                                             "User does not have permission to access or\n"
                                                             "Workbook is opened elsewhere")
                    else:
                        # get appropriate worksheet
                        ws = wb[SHEET_NAME]
                        # get date
                        date = entryCal.get_date()
                        # get row based on date
                        lastItem = searchRowDate(ws, date)
                        # check if empty and modify accordingly
                        modifyRow(ws, lastItem)
                        try:
                            wb.save(filename=file_location)
                        except PermissionError:
                            messagebox.showerror("Fail to save", "Permission Error:\n"
                                                                 "User does not have permission to access or\n"
                                                                 "Workbook is opened elsewhere")
                        except Exception as e:
                            print(e)
                            messagebox.showerror("Fail to save", "something went wrong when saving the changes\n"
                                                                 "your changes has not been submitted\n" + str(e))
                        else:
                            self.resetForm(ents)
                            messagebox.showinfo("Submitted", "Entry successfully submitted")

        # Frames

        # Top Frame
        topFrame = tk.Frame(root)
        topFrame.pack(fill=BOTH, expand=True)

        # Top Left Frame
        leftFrame = tk.Frame(topFrame)
        leftFrame.pack(side=LEFT, fill=BOTH, padx=10, pady=10)

        mainFrame = tk.Frame(leftFrame)
        mainFrame.pack(side=LEFT, fill=BOTH)

        validMainFrame = tk.Frame(leftFrame, pady=40)
        validMainFrame.pack(side=LEFT, fill=BOTH, expand=True)
        validMainFrame.pack_forget()

        # Top Right Frame
        rightFrame = Frame(topFrame)
        rightFrame.pack(expand=True)

        # Dies Frame
        diesFrame = LabelFrame(rightFrame, text='Dies')
        diesFrame.pack(padx=10, pady=10, anchor="w")

        diesMainFrame = Frame(diesFrame)
        diesMainFrame.pack(side=LEFT, fill=BOTH, expand=True)

        validDiesMainFrame = Frame(diesFrame)
        validDiesMainFrame.pack(side=LEFT, fill=BOTH, expand=True)
        validDiesMainFrame.pack_forget()

        # Flex Frame
        flexFrame = LabelFrame(rightFrame, text='Flex')
        flexFrame.pack(padx=10, pady=10, anchor="w")

        flexMainFrame = Frame(flexFrame)
        flexMainFrame.pack(side=LEFT, fill=BOTH, expand=True)

        validFlexMainFrame = Frame(flexFrame)
        validFlexMainFrame.pack(side=LEFT, fill=BOTH, expand=True)
        validFlexMainFrame.pack_forget()

        # THA Frame
        thaFrame = LabelFrame(rightFrame, text='THA')
        thaFrame.pack(padx=10, pady=10, anchor="w")

        thaMainFrame = Frame(thaFrame)
        thaMainFrame.pack(side=LEFT, fill=BOTH, expand=True)

        validThaMainFrame = Frame(thaFrame)
        validThaMainFrame.pack(side=LEFT, fill=BOTH, expand=True)
        validThaMainFrame.pack_forget()

        # bottom frame
        bottomFrame = tk.Frame(root)
        bottomFrame.pack(side=BOTTOM, fill=BOTH, expand=True, padx=10, pady=10)

        # date input
        dateRow = Frame(mainFrame)
        dateRow.pack(side=TOP, fill=X, padx=5, pady=5)
        lblDate = Label(dateRow, text="Date:", anchor='w')
        lblDate.pack(side=LEFT)
        entryCal = DateEntry(dateRow, width=12, background='blue',
                             foreground='white', borderwidth=2)
        entryCal.pack(side=RIGHT, padx=10)

        # making the rest of the form
        ents = {}
        ents1 = self.makeForm(mainFrame, self.fieldsLeft)
        entries2 = ents1

        holderFrame = Frame(mainFrame)
        holderFrame.pack()
        multiFrame = Frame(holderFrame)
        multiFrame.pack(side=LEFT)
        buttonFrame = Frame(holderFrame)
        buttonFrame.pack(side=RIGHT, fill=BOTH)
        self.addField(holderFrame, self.multiWidget)
        Button(buttonFrame, text="+", command=lambda: self.addField(holderFrame, self.multiWidget)).pack(side=BOTTOM,
                                                                                                         anchor='s')
        Button(buttonFrame, text="-", command=lambda: self.delField(self.multiWidget)).pack(side=BOTTOM, anchor='s')

        ents2 = self.makeForm(mainFrame, self.fieldsLeft2)
        entries2.update(ents2)
        ents3 = self.makeForm(diesMainFrame, self.fieldsRight1)
        ents4 = self.makeForm(flexMainFrame, self.fieldsRight2)
        ents5 = self.makeForm(thaMainFrame, self.fieldsRight3)

        ents.update(ents1)
        ents.update(ents2)
        ents.update(ents3)
        ents.update(ents4)
        ents.update(ents5)

        keys = []
        for k in ents:
            keys.append(k)

        keys[len(self.fieldsLeft):len(self.fieldsLeft)] = ["Wafer Lot #", "Wafer ID #"]
        ents.update(self.multiWidget[0])

        print(ents)

        valid = self.makeValidateTxt(validMainFrame, self.fieldsLeft)
        valid.update(self.makeValidateTxt(validMainFrame, self.fieldsLeft2))

        lblYieldTarget = Label(mainFrame, text="yield")
        lblYieldTarget.pack(padx=5, pady=5)

        reg = root.register(self.onValidate)

        # validation to only allow numbers
        ents['Die Bond'].config(textvariable=strQtyIn, validate="key", validatecommand=(reg, '%S'))
        ents['THA Actual Out'].config(textvariable=strQtyOut, validate="key", validatecommand=(reg, '%S'))

        for col in self.fieldsRight1:
            ents[col].config(validate="key", validatecommand=(reg, '%S'))

        for col in self.fieldsRight2:
            ents[col].config(validate="key", validatecommand=(reg, '%S'))

        for col in self.fieldsRight3[0:len(self.fieldsRight3) - 1]:
            ents[col].config(validate="key", validatecommand=(reg, '%S'))

        # buttons
        btnReset = Button(bottomFrame, text="reset", width=12, command=resetForms).pack(side=LEFT, pady=5, padx=5, )
        btnSubmit = Button(bottomFrame, text="submit", width=12, command=handle_submit).pack(side=LEFT, pady=5,
                                                                                             padx=5, )
